from openpyxl import load_workbook
from fastapi import UploadFile, HTTPException
import json
import io
import csv


def parse_file(file: UploadFile) -> dict:
    """
    Parse uploaded Excel or CSV file and return structured data.

    Returns:
        dict with keys: columns (list), rows (list of dicts), row_count (int)
    """
    filename = file.filename.lower()

    if filename.endswith('.csv'):
        return parse_csv_file(file)
    elif filename.endswith(('.xlsx', '.xls')):
        return parse_excel_file(file)
    else:
        raise HTTPException(
            status_code=400,
            detail="Only Excel files (.xlsx, .xls) and CSV files (.csv) are supported"
        )


def parse_csv_file(file: UploadFile) -> dict:
    """Parse uploaded CSV file."""
    try:
        contents = file.file.read().decode('utf-8')
        reader = csv.reader(io.StringIO(contents))

        # Get headers from first row
        headers = next(reader, None)
        if not headers:
            raise HTTPException(status_code=400, detail="CSV file has no column headers")

        # Clean headers
        headers = [h.strip() for h in headers if h.strip()]
        if not headers:
            raise HTTPException(status_code=400, detail="CSV file has no valid column headers")

        # Extract data rows
        rows = []
        for row_idx, row in enumerate(reader, start=1):
            if any(cell.strip() for cell in row):
                row_data = {}
                for i, header in enumerate(headers):
                    value = row[i].strip() if i < len(row) else ""
                    row_data[header] = value

                rows.append({
                    "row_index": row_idx,
                    "content": json.dumps(row_data)
                })

        if not rows:
            raise HTTPException(status_code=400, detail="CSV file has no data rows")

        return {
            "columns": headers,
            "rows": rows,
            "row_count": len(rows)
        }

    except HTTPException:
        raise
    except UnicodeDecodeError:
        raise HTTPException(status_code=400, detail="CSV file must be UTF-8 encoded")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse CSV file: {str(e)}")


def parse_excel_file(file: UploadFile) -> dict:
    """Parse uploaded Excel file."""
    try:
        contents = file.file.read()
        workbook = load_workbook(io.BytesIO(contents), data_only=True)
        sheet = workbook.active

        if sheet is None:
            raise HTTPException(status_code=400, detail="Excel file has no active sheet")

        # Extract headers from first row
        headers = []
        for cell in sheet[1]:
            if cell.value is not None:
                headers.append(str(cell.value))
            else:
                break

        if not headers:
            raise HTTPException(status_code=400, detail="Excel file has no column headers")

        # Extract data rows
        rows = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
            row_values = list(row[:len(headers)])
            if any(cell is not None for cell in row_values):
                row_data = {}
                for i, header in enumerate(headers):
                    value = row_values[i] if i < len(row_values) else None
                    row_data[header] = str(value) if value is not None else ""

                rows.append({
                    "row_index": row_idx,
                    "content": json.dumps(row_data)
                })

        if not rows:
            raise HTTPException(status_code=400, detail="Excel file has no data rows")

        return {
            "columns": headers,
            "rows": rows,
            "row_count": len(rows)
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {str(e)}")
